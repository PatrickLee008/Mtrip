# -*- coding: utf-8 -*-
"""Mtrip 配置信息汇总 Excel 生成脚本
数据来源:deploy/docker-compose.yml、deploy/.env、deploy/openresty/*、
backend/services/*/.env.example、admin-web / merchant-web / supplier-web /
client-app 配置、database/ 初始化与种子 SQL、docs/guides/setup/启动开发指南.md
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"E:\GIT\jiaxu\Mtrip\docs\config\Mtrip配置信息汇总.xlsx"

wb = Workbook()

# ---------- 样式 ----------
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="1668DC")
HEAD_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1668DC")
NOTE_FONT = Font(name="微软雅黑", size=9, color="808080")
WRAP = Alignment(vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def sheet_with_table(ws, title, headers, rows, col_widths, notes=None):
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 26
    # 表头
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[2].height = 22
    # 数据
    for r, row in enumerate(rows, 3):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
            if (r - 3) % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F2F7FD")
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    # 备注
    if notes:
        r = len(rows) + 4
        for i, n in enumerate(notes):
            cell = ws.cell(row=r + i, column=1, value=n)
            cell.font = NOTE_FONT
            ws.merge_cells(start_row=r + i, start_column=1, end_row=r + i, end_column=len(headers))
    ws.freeze_panes = "A3"

# ================= Sheet1 总览 =================
ws = wb.active
ws.title = "总览"
rows = [
    ["项目名称", "Mtrip 海外旅游平台(海外目的地商品/订单/商户/营销一体化平台)"],
    ["技术栈-后端", "PHP 8.1+ / Hyperf 3.1 / Swoole 5.x 微服务(8 个服务),统一由 OpenResty 网关聚合"],
    ["技术栈-前端", "admin-web / merchant-web / supplier-web:Vue3 + Vite5 + Ant Design Vue;client-app:React Native(Expo 51)"],
    ["技术栈-基础设施", "MySQL 8.0(双库)+ Redis 7 + OpenResty 1.25.3.1 网关,统一 Docker Compose 编排"],
    ["统一入口(网关)", "http://localhost:8081(OpenResty,宿主机端口可经 GATEWAY_HOST_PORT 调整)"],
    ["管理后台", "http://localhost:5173(admin-web,Vite dev server,/api 代理至网关 8081)"],
    ["商家中心", "http://localhost:5174(merchant-web,同上代理)"],
    ["供应商中心", "http://localhost:5175(supplier-web,同上代理)"],
    ["移动端", "client-app(Expo),API 地址 http://localhost:8081(真机需改局域网 IP)"],
    ["数据库", "MySQL 双库:mtrip_system(系统域)/ mtrip_business(业务域),宿主机端口 3307"],
    ["缓存", "Redis 7(无密码),宿主机端口 6380,DB 0"],
    ["初始化方式", "Docker 首次启动自动执行 database/ 全部初始化 SQL(建库+DDL+种子);重建需 docker compose down -v"],
    ["一键启动", "cd deploy && cp .env.example .env && docker compose up -d --build"],
    ["默认管理账号", "admin / Admin@123456(mtrip_system.sys_admin,超管,首次登录后须修改)"],
    ["默认数据库账号", "mtrip / mtrip@2026(root 密码:root@2026)"],
    ["配置事实源", "deploy/docker-compose.yml、deploy/.env、backend/services/*/.env.example、各前端 .env*、database/*"],
]
sheet_with_table(
    ws, "Mtrip 项目配置总览",
    ["项目", "说明"],
    rows,
    [22, 100],
    notes=["说明:本文档由脚本自动汇总自仓库内配置文件,更新时间 2026-08-23;端口/密钥调整后请同步更新。"],
)

# ================= Sheet2 后端微服务 =================
ws = wb.create_sheet("后端微服务")
rows = [
    ["system-service", "9501", "mtrip_system(主)+ mtrip_business(共享 system 连接)", "系统域:管理员登录/RBAC/菜单/站点/全局配置/客户端密钥/审计日志/帮助中心/特性开关", "auth、config、sys、help、site、client(admin);site、theme(app)"],
    ["user-service", "9502", "mtrip_business", "C 端用户/会员/出行人/收藏/推荐/通知/聊天/黑名单", "user(admin);auth、user、chat(app)"],
    ["goods-service", "9503", "mtrip_business", "商品/类目/价格(双重价格)/评价/筛选", "goods(admin);goods(app);goods(merchant)"],
    ["order-service", "9504", "mtrip_business", "订单/预订/核销/行程", "order(admin);order(app);order(merchant)"],
    ["merchant-service", "9505", "mtrip_business", "商户/供应商/集团门店/入驻验证/RBAC/合规/排名", "merchant、supplier、compliance(admin);auth、account、role、store、goods、settle(merchant/supplier)"],
    ["finance-service", "9506", "mtrip_business", "财务/结算/账户分录", "finance(admin)"],
    ["marketing-service", "9507", "mtrip_business", "营销/优惠券/促销/长住/达人联盟/出资分账", "marketing、affiliate(admin);marketing(app)"],
    ["payment-service", "9508", "mtrip_business", "支付渠道抽象与回调(无管理端路由)", "/api/v1/payment/callback/{channel}(回调专用)"],
]
sheet_with_table(
    ws, "后端微服务清单(Hyperf 3.1,容器端口=宿主机端口)",
    ["服务名(Compose)", "端口", "依赖数据库", "职责", "网关路由前缀→模块"],
    rows,
    [20, 8, 30, 52, 46],
    notes=[
        "健康检查:GET http://127.0.0.1:{port}/healthz",
        "各服务均注册共享包全局中间件(CORS/JWT/客户端签名/传输加密/提交防重),JWT 密钥与 AES 密钥全平台必须一致。",
        "开发期热更新:deploy/docker-compose.override.yml 挂载源码,改代码 docker compose restart xxx-service 约 2 秒生效。",
    ],
)

# ================= Sheet3 基础设施 =================
ws = wb.create_sheet("基础设施")
rows = [
    ["MySQL 8.0", "mysql(容器)", "3307 → 3306", "mtrip_system / mtrip_business(字符集 utf8mb4,排序 utf8mb4_bin,时区 +00:00)", "root / root@2026;mtrip / mtrip@2026(仅两库 ALL 权限)", "healthcheck: mysqladmin ping;首次启动自动执行 database/ 初始化 SQL"],
    ["Redis 7-alpine", "redis(容器)", "6380 → 6379", "DB 0,无密码(REDIS_AUTH 空)", "—", "healthcheck: redis-cli ping;用于缓存/登录失败锁定/签名 nonce 去重/并发锁"],
    ["OpenResty 网关", "gateway(容器)", "8081 → 80", "统一 API 分发 /api/v1/{admin|app|merchant|supplier}/*;IP 限流 30r/s 突发 60;client_max_body_size 20m", "—", "探活:http://localhost:8081/healthz;配置挂载 deploy/openresty/"],
    ["Docker Compose", "项目名 mtrip", "—", "编排 8 微服务 + MySQL + Redis + 网关;override 文件挂载源码与日志(deploy/logs/<服务名>/)", "—", "启动:cd deploy;docker compose up -d --build"],
]
sheet_with_table(
    ws, "基础设施与中间件",
    ["组件", "Compose 服务名", "端口(宿主机→容器)", "说明", "账号", "备注"],
    rows,
    [16, 16, 18, 56, 34, 44],
    notes=["宿主机端口可通过 deploy/.env 调整:MYSQL_HOST_PORT(默认 3307)、REDIS_HOST_PORT(默认 6380)、GATEWAY_HOST_PORT(默认 8081)。"],
)

# ================= Sheet4 前端应用 =================
ws = wb.create_sheet("前端应用")
rows = [
    ["admin-web", "Mtrip 海外旅游平台管理后台", "Vue3 + Vite5 + Ant Design Vue + ECharts", "http://localhost:5173", "/api → http://127.0.0.1:8081(vite.config.ts)", "VITE_API_BASE=/api/v1;VITE_APP_TITLE=Mtrip 海外旅游平台管理后台;VITE_LOGIN_AES_KEY=mtrip-dev-admin-login-key-change-me(与后端 MTRIP_ADMIN_AES_KEY 同值)", "启动:npm run dev;构建验收:npm run build 零 TS 报错"],
    ["merchant-web", "商家中心", "Vue3 + Vite5 + Ant Design Vue", "http://localhost:5174", "/api → http://127.0.0.1:8081", "VITE_API_BASE=/api/v1;VITE_APP_TITLE=商家中心;VITE_LOGIN_AES_KEY=(留空走明文,生产与后端 MTRIP_MERCHANT_AES_KEY 同值)", "启动:npm run dev;API 前缀 /api/v1/merchant/*"],
    ["supplier-web", "供应商中心", "Vue3 + Vite5 + Ant Design Vue", "http://localhost:5175", "/api → http://127.0.0.1:8081", "VITE_API_BASE=/api/v1;VITE_APP_TITLE=供应商中心;VITE_LOGIN_AES_KEY=(留空走明文,生产与后端 MTRIP_SUPPLIER_AES_KEY 同值)", "启动:npm run dev;API 前缀 /api/v1/supplier/*"],
    ["client-app", "Mtrip 移动端(Expo)", "React Native + Expo 51(slug: mtrip-client-app)", "Expo DevTools(npm start;按 a/i/w 启动 Android/iOS/Web)", "直连网关 http://localhost:8081(真机需改局域网 IP)", "EXPO_PUBLIC_API_BASE_URL=http://localhost:8081;EXPO_PUBLIC_DEFAULT_SITE_ID=1;EXPO_PUBLIC_CLIENT_ID / EXPO_PUBLIC_CLIENT_SECRET(后台客户端管理创建后填入)", "scheme: mtrip;bundleId/package: com.mtrip.clientapp;签名 HMAC-SHA256 + AES 加密登录"],
]
sheet_with_table(
    ws, "前端应用清单(4 个端)",
    ["应用目录", "应用名称", "技术栈", "访问地址/入口", "API 代理与目标", "关键环境变量", "启动与说明"],
    rows,
    [14, 24, 30, 34, 34, 56, 44],
    notes=[
        "四个端开发态均经网关 8081 访问后端;未启动网关时可临时将代理指向单个服务端口(如 9503)逐服务联调。",
        "改 .env 后需重启对应 dev server / Expo(Expo 需 npm start -- --clear)。",
    ],
)

# ================= Sheet5 账号与密钥 =================
ws = wb.create_sheet("账号与密钥")
rows = [
    ["管理后台超管账号", "admin / Admin@123456", "mtrip_system.sys_admin(id=1,is_super=1,site_id=0)", "database/seed/01-admin-role.sql", "bcrypt 存储;首次登录后立即修改;超管代码侧跳过权限校验"],
    ["MySQL root", "root / root@2026", "MySQL 实例", "deploy/.env(MYSQL_ROOT_PASSWORD)", "仅容器/宿主机 3307 连接使用"],
    ["MySQL 应用账号", "mtrip / mtrip@2026", "mtrip_system、mtrip_business 两库 ALL 权限", "database/init/00-create-databases.sql + deploy/.env", "修改密码需同步改 SQL 脚本或手动 ALTER USER"],
    ["移动端客户端密钥", "ClientId / ClientSecret(经后台创建,无预置种子)", "mtrip_system.sys_client(Secret AES 加密存储)", "database/system/06-client.sql", "admin 后台「配置→客户端管理」创建;Secret 仅创建/重置时展示一次;填入 client-app .env"],
    ["JWT 密钥", "mtrip-dev-jwt-secret-change-me", "全平台统一,跨服务验签", "deploy/.env(MTRIP_JWT_SECRET)", "生产必须换强随机值;TTL:管理端 7200s,App 604800s"],
    ["通用 AES 密钥", "mtrip-dev-aes-key-change-me", "传输加密(登录/注册 AES-256-CBC)", "deploy/.env(MTRIP_AES_KEY)", "生产必须换强随机值"],
    ["管理端登录加密密钥", "mtrip-dev-admin-login-key-change-me", "admin 登录传输加密", "deploy/.env(MTRIP_ADMIN_AES_KEY) = admin-web VITE_LOGIN_AES_KEY", "生产换强随机串,两端保持同值"],
    ["站点种子", "全球(1)→欧洲(2)→法国(3)→巴黎(4)", "mtrip_system.sys_site 示例站点树", "database/seed/03-config-site.sql", "巴黎站(id=4)含 VAT 20%/城市税/佣金 10% 差异化配置;客服邮箱 support@mtrip.com"],
]
sheet_with_table(
    ws, "账号与密钥清单(默认开发值,生产必须更换)",
    ["项目", "值(账号/密钥)", "作用范围", "配置来源", "说明"],
    rows,
    [20, 44, 34, 40, 50],
    notes=["安全提醒:以上均为开发默认值;生产环境 JWT/AES/登录密钥必须替换为强随机值,默认账号密码必须修改。"],
)

# ================= Sheet6 网关路由表 =================
ws = wb.create_sheet("网关路由表")
rows = [
    ["admin(管理后台)", "/api/v1/admin/auth|config|sys|help|site|client", "system-service:9501"],
    ["admin(管理后台)", "/api/v1/admin/merchant|supplier|compliance", "merchant-service:9505"],
    ["admin(管理后台)", "/api/v1/admin/goods", "goods-service:9503"],
    ["admin(管理后台)", "/api/v1/admin/order", "order-service:9504"],
    ["admin(管理后台)", "/api/v1/admin/finance", "finance-service:9506"],
    ["admin(管理后台)", "/api/v1/admin/user|chat", "user-service:9502"],
    ["admin(管理后台)", "/api/v1/admin/marketing|affiliate", "marketing-service:9507"],
    ["app(移动端)", "/api/v1/app/site|theme", "system-service:9501"],
    ["app(移动端)", "/api/v1/app/auth|user|chat", "user-service:9502"],
    ["app(移动端)", "/api/v1/app/goods", "goods-service:9503"],
    ["app(移动端)", "/api/v1/app/order", "order-service:9504"],
    ["app(移动端)", "/api/v1/app/marketing", "marketing-service:9507"],
    ["merchant(商家端)", "/api/v1/merchant/auth|account|role|store", "merchant-service:9505"],
    ["merchant(商家端)", "/api/v1/merchant/order", "order-service:9504"],
    ["merchant(商家端)", "/api/v1/merchant/goods", "goods-service:9503"],
    ["supplier(供应商端)", "/api/v1/supplier/auth|account|role|goods|settle", "merchant-service:9505"],
    ["payment(支付回调)", "/api/v1/payment/callback/{channel}", "payment-service:9508"],
    ["探活", "/healthz", "网关本机直答 {code:0}"],
]
sheet_with_table(
    ws, "OpenResty 网关路由表(deploy/openresty/conf.d/mtrip.conf)",
    ["路由组(客户端)", "路径前缀(/api/v1/...)", "上游服务"],
    rows,
    [22, 62, 30],
    notes=[
        "限流:按客户端 IP 30r/s,突发 60,超限返回 429(code 42900);上游异常统一 JSON 502(code 50200)。",
        "CORS 预检 OPTIONS 由网关直接应答 204;后端 CorsMiddleware 重复头被 proxy_hide_header 剥除。",
        "新增自定义请求头需同步扩充 Allow-Headers 并 docker exec mtrip-gateway-1 nginx -s reload。",
    ],
)

# ================= Sheet7 环境变量 =================
ws = wb.create_sheet("环境变量")
rows = [
    ["APP_ENV", "dev", "运行环境"],
    ["SCAN_CACHEABLE", "false", "Hyperf 注解缓存;开发 false(改注解重启即生效),生产必须 true"],
    ["MAX_REQUEST", "10000", "worker 处理 N 个请求后平滑重建,防内存泄漏"],
    ["MYSQL_ROOT_PASSWORD", "root@2026", "MySQL root 密码"],
    ["MYSQL_USER / MYSQL_PASSWORD", "mtrip / mtrip@2026", "MySQL 应用账号(与 00-create-databases.sql 同步)"],
    ["MYSQL_HOST_PORT", "3307", "MySQL 宿主机映射端口(容器内 3306)"],
    ["REDIS_HOST_PORT", "6380", "Redis 宿主机映射端口(容器内 6379)"],
    ["GATEWAY_HOST_PORT", "8081", "网关宿主机映射端口(容器内 80,8080 易被占用故默认 8081)"],
    ["MTRIP_JWT_SECRET", "mtrip-dev-jwt-secret-change-me", "全平台统一 JWT 密钥,跨服务验签,生产必须更换"],
    ["MTRIP_JWT_TTL / MTRIP_APP_JWT_TTL", "7200 / 604800", "管理端 / App 端 Token 有效期(秒)"],
    ["MTRIP_AES_KEY", "mtrip-dev-aes-key-change-me", "通用 AES 传输加密密钥,生产必须更换"],
    ["MTRIP_ADMIN_AES_KEY", "mtrip-dev-admin-login-key-change-me", "管理端登录加密密钥,与 admin-web VITE_LOGIN_AES_KEY 同值"],
    ["MTRIP_CLIENT_SIGN", "true", "app 接口 HMAC-SHA256 签名校验开关,生产必须 true;本地 curl 可临时 false"],
    ["MTRIP_SIGN_TTL", "3600", "签名时间窗口(秒),超窗报请求已过期"],
    ["MTRIP_PAYLOAD_ENCRYPT", "true", "登录/注册强制 AES 加密传输开关,生产必须 true"],
    ["MTRIP_REQUEST_LOG", "true(.env)/false(默认)", "全量请求日志排查模式,写 deploy/logs/<服务名>/request-*.log,生产保持 false"],
    ["MTRIP_SUBMIT_LOCK", "true", "写操作 Redis 并发锁开关(重复提交拒绝 42902)"],
    ["MTRIP_SUBMIT_LOCK_TTL", "10", "并发锁兜底 TTL(秒)"],
    ["MTRIP_FORM_ID_TTL", "300", "X-Form-Id 幂等窗口(秒)"],
    ["WORKER_NUM", "2", "每个服务 Swoole worker 数"],
    ["DB_MAX_CONNECTIONS", "10", "每个服务数据库连接池上限"],
]
sheet_with_table(
    ws, "环境变量清单(主文件 deploy/.env;微服务另有 backend/services/*/.env.example 手动方式用)",
    ["变量", "默认值", "说明"],
    rows,
    [34, 36, 66],
    notes=[
        "docker 环境改 .env 后需 docker compose up -d 重建容器生效;手动方式各服务 .env 的 MTRIP_JWT_SECRET/MTRIP_AES_KEY 必须一致。",
        "后端 .env.example 额外含:MTRIP_LOGIN_FAIL_LIMIT=5(登录失败锁定次数)、MTRIP_LOGIN_LOCK_MINUTES=30(锁定时长)。",
    ],
)

# ================= Sheet8 数据库初始化 =================
ws = wb.create_sheet("数据库初始化脚本")
rows = [
    ["00", "database/init/00-create-databases.sql", "mtrip_system / mtrip_business", "建库(utf8mb4_bin)+ 创建 mtrip 应用账号授权"],
    ["10-15", "database/system/01~06(admin-role-menu/log/config-site/storage-file/pay-sms-map/client)", "mtrip_system", "系统域 DDL:管理员/RBAC/菜单/站点配置/存储/支付短信映射/客户端密钥"],
    ["20-30", "database/merchant/01~10(merchant/supplier/group-store/admin-account-type/rbac/verify-workflow/application/account-security…)", "mtrip_business", "商户/供应商域 DDL:商户/供应商/集团门店/账号类型/RBAC/验证工作流/入驻流水线/账户安全(2FA)"],
    ["31-39c", "database/merchant/11~21 + goods/01(notify/confirmation/impersonation/ranking/verify-rework/onboarding 队列与阶段 KYC 等)", "mtrip_business", "商户整改各阶段 DDL:通知中心/KYC/代入会话/排名/五队列入驻/业务单元 KYC/四节点阶段回归"],
    ["40-41", "database/order/01-order.sql、02-verify.sql", "mtrip_business", "订单/核销 DDL"],
    ["50", "database/finance/01-finance.sql", "mtrip_business", "财务/结算 DDL"],
    ["60", "database/marketing/01-marketing.sql", "mtrip_business", "营销 DDL"],
    ["70-71", "database/user/01-user.sql、02-user-app.sql", "mtrip_business", "C 端用户 DDL"],
    ["80-99f", "database/{goods,order,marketing,user,system,finance}/consumer-*、affiliate、enduser-blacklist、helpcenter、compliance、promotion-extras、feature-flag 等增量迁移", "mtrip_business / mtrip_system", "Consumer App PRD v1.0 增量(双重价格/预订/长住/出行人/评价/收藏/推荐/通知/主题/申诉/分账/聊天/行程/筛选/活动)、达人联盟、黑名单、帮助中心、合规、促销、特性开关"],
    ["90-92", "database/seed/01-admin-role.sql、02-menu.sql、03-config-site.sql", "mtrip_system", "种子:超管账号(admin/Admin@123456)+超管角色、管理后台菜单与 perm_key、全局配置+示例站点树(全球→欧洲→法国→巴黎)"],
    ["26-27", "database/seed/04-merchant-menu.sql、05-supplier-menu.sql", "mtrip_business", "种子:商家端/供应商端菜单与 RBAC"],
]
sheet_with_table(
    ws, "数据库初始化脚本(docker-entrypoint-initdb.d 按文件名顺序执行,共 60+ SQL)",
    ["执行序号(容器内)", "脚本来源", "目标库", "内容"],
    rows,
    [18, 72, 24, 64],
    notes=[
        "首次启动自动执行;已启动过的库不会重跑 init SQL,增量 DDL 需手动进 3307 执行或 docker compose down -v 重建(会清数据)。",
        "手动初始化顺序见 docs/guides/setup/启动开发指南.md 第 3 节(先 init → system → business → seed)。",
        "种子菜单 perm_key 与后端 #[Permission('...')] 注解、前端 v-perm 指令三方对齐(项目硬约定)。",
    ],
)

wb.save(OUT)
print("saved:", OUT)
